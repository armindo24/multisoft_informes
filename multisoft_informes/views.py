from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission
from django.contrib.sessions.models import Session
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from datetime import datetime, date
import json
import requests
import custom_permissions
import xlrd
from openpyxl import load_workbook

NODE_API_TIMEOUT = 20
NODE_API_TIMEOUT_LONG = 600
NODE_API_BASE = 'http://localhost:3000/api/v1/'
ASIENTOS_MIGRATION_BATCH_SIZE = 1


ASIENTOS_REQUIRED_COLUMNS = [
    'cod_empresa',
    'tipoasiento',
    'nrocompr',
    'nrotransaccion',
    'linea',
    'codplancta',
    'concepto',
    'dbcr',
    'importe',
    'importeme',
    'fecha',
]

ASIENTOS_HEADER_ALIASES = {
    'codempresa': 'cod_empresa',
    'cod_empresa': 'cod_empresa',
    'tipoasiento': 'tipoasiento',
    'nrocompr': 'nrocompr',
    'nrocomprobante': 'nrocompr',
    'nrotransaccion': 'nrotransaccion',
    'nrotransac': 'nrotransaccion',
    'line': 'linea',
    'linea': 'linea',
    'lin': 'linea',
    'codplancta': 'codplancta',
    'codplancuenta': 'codplancta',
    'codplanaux': 'codplanaux',
    'codplanau': 'codplanaux',
    'concepto': 'concepto',
    'dbcr': 'dbcr',
    'importe': 'importe',
    'importeme': 'importeme',
    'fecha': 'fecha',
}

VENTAS_REQUIRED_COLUMNS = [
    'fha_cbte',
    'linea',
    'comp_numero',
    'razon_social',
    'ruc',
    'cta_cliente',
    'aux_cliente',
    'cod_articulo',
    'descrip',
    'cantidad',
    'pr_unit',
    'total',
    'cod_con_vta',
    'codmoneda',
    'fact_cambio',
    'direccion',
    'telefono',
    'observ',
    'to_gravado_5',
    'total_iva_5',
    'to_gravado_10',
    'total_iva_10',
    'exenta',
    'anulado',
    'anulado_el',
    'anulado_por',
    'total_neto',
    'porcpartgravado',
    'montopartgravado',
    'nrotimb',
    'cod_establecimiento',
    'cod_ptoexpedicion',
    'cuotas',
]

VENTAS_HEADER_ALIASES = {
    'fhacbte': 'fha_cbte',
    'fha_cbte': 'fha_cbte',
    'fecha': 'fha_cbte',
    'linea': 'linea',
    'lineas': 'linea',
    'compnumero': 'comp_numero',
    'comp_numero': 'comp_numero',
    'compnro': 'comp_numero',
    'razonsocial': 'razon_social',
    'razon_social': 'razon_social',
    'ruc': 'ruc',
    'ctacliente': 'cta_cliente',
    'cta_cliente': 'cta_cliente',
    'auxcliente': 'aux_cliente',
    'aux_cliente': 'aux_cliente',
    'codarticulo': 'cod_articulo',
    'cod_articulo': 'cod_articulo',
    'descrip': 'descrip',
    'cantidad': 'cantidad',
    'prunit': 'pr_unit',
    'pr_unit': 'pr_unit',
    'total': 'total',
    'codconvta': 'cod_con_vta',
    'cod_con_vta': 'cod_con_vta',
    'codmoneda': 'codmoneda',
    'factcambio': 'fact_cambio',
    'fact_cambio': 'fact_cambio',
    'direccion': 'direccion',
    'telefono': 'telefono',
    'observ': 'observ',
    'togravado5': 'to_gravado_5',
    'to_gravado_5': 'to_gravado_5',
    'totaliva5': 'total_iva_5',
    'total_iva_5': 'total_iva_5',
    'togravado10': 'to_gravado_10',
    'to_gravado_10': 'to_gravado_10',
    'totaliva10': 'total_iva_10',
    'total_iva_10': 'total_iva_10',
    'exenta': 'exenta',
    'anulado': 'anulado',
    'anuladoel': 'anulado_el',
    'anulado_el': 'anulado_el',
    'anuladopor': 'anulado_por',
    'anulado_por': 'anulado_por',
    'totalneto': 'total_neto',
    'total_neto': 'total_neto',
    'porcpartgravado': 'porcpartgravado',
    'montopartgravado': 'montopartgravado',
    'nrotimb': 'nrotimb',
    'codestablecimiento': 'cod_establecimiento',
    'cod_establecimiento': 'cod_establecimiento',
    'codptoexpedicion': 'cod_ptoexpedicion',
    'cod_ptoexpedicion': 'cod_ptoexpedicion',
    'cuotas': 'cuotas',
}


def _normalize_excel_header(value, aliases=None):
    raw = str(value or '').strip().lower()
    raw = raw.replace(' ', '').replace('-', '').replace('.', '').replace('/', '')
    aliases = aliases or ASIENTOS_HEADER_ALIASES
    return aliases.get(raw, raw)


def _chunk_asientos_rows(rows, batch_size=ASIENTOS_MIGRATION_BATCH_SIZE):
    grouped = {}
    order = []
    for row in rows or []:
        key = (
            str(row.get('cod_empresa') or ''),
            str(row.get('nrotransaccion') or ''),
        )
        if key not in grouped:
            grouped[key] = []
            order.append(key)
        grouped[key].append(row)

    batches = []
    current = []
    count_groups = 0
    for key in order:
        current.extend(grouped[key])
        count_groups += 1
        if count_groups >= batch_size:
            batches.append(current)
            current = []
            count_groups = 0
    if current:
        batches.append(current)
    return batches


def _normalize_string(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_ruc(value):
    ruc = _normalize_string(value)
    if not ruc:
        return ''
    cleaned = ruc.replace(' ', '').replace('.', '').replace('_', '').replace('/', '')
    if '-' not in cleaned and cleaned.isdigit() and len(cleaned) > 1:
        return cleaned[:-1] + '-' + cleaned[-1]
    return cleaned


def _normalize_number(value):
    if value in (None, ''):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip().replace('.', '').replace(',', '.')
    try:
        return float(txt)
    except Exception:
        return None


def _normalize_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    txt = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(txt, fmt).date()
        except Exception:
            continue
    return None


def _normalize_int(value):
    if value in (None, ''):
        return None
    try:
        return int(float(value))
    except Exception:
        try:
            return int(str(value).strip())
        except Exception:
            return None


def _is_integer_like(value):
    if value in (None, ''):
        return False
    try:
        return float(value).is_integer()
    except Exception:
        try:
            int(str(value).strip())
            return True
        except Exception:
            return False


def _read_asientos_excel(uploaded_file):
    name = (uploaded_file.name or '').lower()
    if name.endswith('.xlsx'):
        wb = load_workbook(uploaded_file, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))
    if name.endswith('.xls'):
        book = xlrd.open_workbook(file_contents=uploaded_file.read())
        sheet = book.sheet_by_index(0)
        rows = []
        for rx in range(sheet.nrows):
            row = []
            for cx in range(sheet.ncols):
                cell = sheet.cell(rx, cx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, book.datemode)
                row.append(value)
            rows.append(tuple(row))
        return rows
    raise ValueError('Formato no soportado. Use un archivo .xls o .xlsx')


def _node_get(path, params=None, timeout=None):
    return requests.get(
        NODE_API_BASE + path.lstrip('/'),
        params=params or {},
        timeout=timeout or NODE_API_TIMEOUT
    )


def _node_post(path, json=None, timeout=None):
    last_exc = None
    for _ in range(2):
        try:
            return requests.post(
                NODE_API_BASE + path.lstrip('/'),
                json=json or {},
                timeout=timeout or NODE_API_TIMEOUT
            )
        except requests.exceptions.ReadTimeout as exc:
            last_exc = exc
            continue
    raise last_exc


def _load_asientos_reference_data(empresa, periodo, rows):
    if not empresa or not periodo:
        return {'plancta': {}, 'planauxi': {}, 'tipoasiento': {}, 'control': None}

    plancta_resp = _node_get('cuenta/plancta/{0}/{1}'.format(empresa, periodo), timeout=NODE_API_TIMEOUT_LONG)
    plancta_resp.raise_for_status()
    plancta_rows = (plancta_resp.json() or {}).get('data', [])
    plancta_map = {}
    for row in plancta_rows:
        cod = _normalize_string(row.get('codplancta') if isinstance(row, dict) else None)
        if not cod:
            cod = _normalize_string(row.get('CodPlanCta') if isinstance(row, dict) else None)
        if cod:
            plancta_map[cod] = row

    codplancta_values = []
    for row in rows[1:]:
        if not any(str(v or '').strip() for v in row):
            continue
        if len(row) > 5:
            codplancta = _normalize_string(row[5])
            if codplancta:
                codplancta_values.append(codplancta)

    planaux_map = {}
    if codplancta_values:
        cuentad = min(codplancta_values)
        cuentah = max(codplancta_values)
        aux_resp = _node_post(
            'cuentaauxi/query/{0}/{1}/{2}/{3}'.format(periodo, empresa, cuentad, cuentah),
            json={'nombre': '', 'codigo': ''},
            timeout=NODE_API_TIMEOUT_LONG
        )
        aux_resp.raise_for_status()
        aux_rows = (aux_resp.json() or {}).get('data', [])
        for row in aux_rows:
            cta = _normalize_string(row.get('CodPlanCta') if isinstance(row, dict) else None)
            aux = _normalize_string(row.get('CodPlanAux') if isinstance(row, dict) else None)
            if cta and aux:
                planaux_map[(cta, aux)] = row

    tipo_resp = _node_get('tipoasiento/select/', timeout=NODE_API_TIMEOUT_LONG)
    tipo_resp.raise_for_status()
    tipo_rows = (tipo_resp.json() or {}).get('data', [])
    tipo_map = {}
    for row in tipo_rows:
        cod = _normalize_string(row.get('TipoAsiento') if isinstance(row, dict) else None)
        if cod:
            tipo_map[cod] = row

    control_resp = _node_get('control/cierre/{0}/{1}'.format(empresa, periodo), timeout=NODE_API_TIMEOUT_LONG)
    control_resp.raise_for_status()
    control_rows = (control_resp.json() or {}).get('data', [])
    control_row = control_rows[0] if control_rows else None

    return {
        'plancta': plancta_map,
        'planauxi': planaux_map,
        'tipoasiento': tipo_map,
        'control': control_row,
    }


def _load_ventas_reference_data(empresa, periodo, rows, tipo_comprobante='', sucursal=''):
    if not empresa or not periodo:
        return {
            'plancta': {},
            'planauxi': {},
            'articulos': set(),
            'condiciones': set(),
            'monedas': set(),
            'clientes': {},
            'created_clientes': set(),
            'missing_clientes': set(),
            'control': None,
        }

    cuentas = []
    auxiliares = []
    articulos = []
    condiciones = []
    monedas = []
    rucs = []
    for row in rows[1:]:
        if not any(str(v or '').strip() for v in row):
            continue
        if len(row) > 5:
            cta = _normalize_string(row[5])
            if cta:
                cuentas.append(cta)
        if len(row) > 6:
            aux = _normalize_string(row[6])
            if aux:
                auxiliares.append(aux)
        if len(row) > 7:
            art = _normalize_string(row[7])
            if art:
                articulos.append(art)
        if len(row) > 12:
            cond = _normalize_string(row[12])
            if cond:
                condiciones.append(cond)
        if len(row) > 13:
            mon = _normalize_string(row[13])
            if mon:
                monedas.append(mon.upper())
        if len(row) > 4:
            ruc = _normalize_ruc(row[4])
            if ruc:
                rucs.append(ruc)

    refs_resp = _node_post(
        'migraciones/ventas/referencias',
        json={
            'empresa': empresa,
            'periodo': periodo,
            'cuentas_cliente': sorted(set(cuentas)),
            'aux_clientes': sorted(set(auxiliares)),
            'articulos': sorted(set(articulos)),
            'condiciones': sorted(set(condiciones)),
            'monedas': sorted(set(monedas)),
            'rucs': sorted(set(rucs)),
        },
        timeout=NODE_API_TIMEOUT_LONG
    )
    refs_resp.raise_for_status()
    refs_json = refs_resp.json() or {}
    refs_data = refs_json.get('data', {}) if isinstance(refs_json, dict) else {}

    plancta_map = {}
    for row in refs_data.get('plancta', []) or []:
        cod = _normalize_string(row.get('codplancta', row.get('CodPlanCta', '')) if isinstance(row, dict) else '')
        if cod:
            plancta_map[cod] = row

    planaux_map = {}
    for row in refs_data.get('planauxi', []) or []:
        cta = _normalize_string(row.get('codplancta', row.get('CodPlanCta', '')) if isinstance(row, dict) else '')
        aux = _normalize_string(row.get('codplanaux', row.get('CodPlanAux', '')) if isinstance(row, dict) else '')
        if cta and aux:
            planaux_map[(cta, aux)] = row

    articulo_set = set()
    for row in refs_data.get('articulos', []) or []:
        cod = _normalize_string(row.get('cod_articulo', row.get('Cod_Articulo', '')) if isinstance(row, dict) else '')
        if cod:
            articulo_set.add(cod)

    condiciones_set = set()
    for row in refs_data.get('terminos', []) or []:
        cod = _normalize_string(row.get('cod_con_vta', row.get('Cod_Con_Vta', '')) if isinstance(row, dict) else '')
        if cod:
            condiciones_set.add(cod)

    monedas_set = set()
    for row in refs_data.get('monedas', []) or []:
        cod = _normalize_string(row.get('simbolo', row.get('Simbolo', '')) if isinstance(row, dict) else '')
        if cod:
            monedas_set.add(cod.upper())

    clientes_map = {}
    for row in refs_data.get('clientes', []) or []:
        if not isinstance(row, dict):
            continue
        ruc = _normalize_ruc(row.get('RUC', row.get('ruc', '')))
        cod = _normalize_string(row.get('Cod_Cliente', row.get('cod_cliente', '')))
        if ruc:
            clientes_map[ruc] = cod

    created_clientes = set()
    missing_clientes = set()
    missing_payload = []
    seen_missing = set()
    for row in rows[1:]:
        if not any(str(v or '').strip() for v in row):
            continue
        ruc = _normalize_ruc(row[4]) if len(row) > 4 else ''
        razon = _normalize_string(row[3]) if len(row) > 3 else ''
        cond = _normalize_string(row[12]) if len(row) > 12 else ''
        mon = _normalize_string(row[13]) if len(row) > 13 else ''
        cta = _normalize_string(row[5]) if len(row) > 5 else ''
        aux = _normalize_string(row[6]) if len(row) > 6 else ''
        if not ruc:
            ruc = '44444401-7'
        if not razon:
            razon = 'Sin Nombre'
        if ruc and ruc not in clientes_map and ruc not in seen_missing:
            missing_payload.append({
                'ruc': ruc,
                'razon_social': razon,
                'cod_con_vta': cond,
                'codmoneda': mon,
                'codplancta': cta,
                'codplanaux': aux,
            })
            seen_missing.add(ruc)

    if missing_payload:
        for item in missing_payload:
            r = _normalize_ruc(item.get('ruc'))
            if r:
                missing_clientes.add(r)
        try:
            create_resp = _node_post(
                'migraciones/ventas/clientes/crear',
                json={
                    'empresa': empresa,
                    'cod_tp_comp': tipo_comprobante,
                    'cod_sucursal': sucursal,
                    'cod_usuario': '',
                    'clientes': missing_payload,
                },
                timeout=NODE_API_TIMEOUT_LONG
            )
            create_resp.raise_for_status()
            create_json = create_resp.json() or {}
            for row in create_json.get('created', []) or []:
                ruc = _normalize_ruc(row.get('ruc', ''))
                cod = _normalize_string(row.get('cod_cliente', ''))
                if ruc and cod:
                    clientes_map[ruc] = cod
                    created_clientes.add(ruc)
        except Exception:
            pass
        if missing_clientes:
            try:
                refresh_resp = _node_post(
                    'migraciones/ventas/referencias',
                    json={
                        'empresa': empresa,
                        'periodo': periodo,
                        'cuentas_cliente': [],
                        'aux_clientes': [],
                        'articulos': [],
                        'condiciones': [],
                        'monedas': [],
                        'rucs': sorted(missing_clientes),
                    },
                    timeout=NODE_API_TIMEOUT_LONG
                )
                refresh_resp.raise_for_status()
                refresh_json = refresh_resp.json() or {}
                refresh_data = refresh_json.get('data', {}) if isinstance(refresh_json, dict) else {}
                for row in refresh_data.get('clientes', []) or []:
                    if not isinstance(row, dict):
                        continue
                    ruc = _normalize_ruc(row.get('RUC', row.get('ruc', '')))
                    cod = _normalize_string(row.get('Cod_Cliente', row.get('cod_cliente', '')))
                    if ruc and cod:
                        clientes_map[ruc] = cod
                missing_clientes = set([r for r in missing_clientes if r not in clientes_map])
            except Exception:
                pass

    control_row = None
    try:
        control_resp = _node_get('control/cierre/{0}/{1}'.format(empresa, periodo), timeout=NODE_API_TIMEOUT_LONG)
        control_resp.raise_for_status()
        control_rows = (control_resp.json() or {}).get('data', [])
        control_row = control_rows[0] if control_rows else None
    except Exception:
        control_row = None

    return {
        'plancta': plancta_map,
        'planauxi': planaux_map,
        'articulos': articulo_set,
        'condiciones': condiciones_set,
        'monedas': monedas_set,
        'clientes': clientes_map,
        'created_clientes': created_clientes,
        'missing_clientes': missing_clientes,
        'control': control_row,
    }


def _validate_asientos_rows(rows, empresa, periodo):
    if not rows:
        return {
            'headers': [],
            'preview_rows': [],
            'errors': ['El archivo no contiene filas.'],
            'summary': {'total_rows': 0, 'valid_rows': 0, 'error_rows': 0}
        }

    headers = [_normalize_excel_header(h) for h in list(rows[0])]
    missing = [col for col in ASIENTOS_REQUIRED_COLUMNS if col not in headers]
    if missing:
        return {
            'headers': headers,
            'preview_rows': [],
            'errors': ['Faltan columnas obligatorias: ' + ', '.join(missing)],
            'summary': {'total_rows': 0, 'valid_rows': 0, 'error_rows': 0}
        }

    header_index = {header: idx for idx, header in enumerate(headers)}
    preview_rows = []
    errors = []
    total_rows = 0
    processed_rows = []
    references = _load_asientos_reference_data(empresa, periodo, rows)
    plancta_map = references.get('plancta', {})
    planaux_map = references.get('planauxi', {})
    tipoasiento_map = references.get('tipoasiento', {})
    control_row = references.get('control')

    for rownum, row in enumerate(rows[1:], start=2):
        if not any(str(v or '').strip() for v in row):
            continue
        total_rows += 1

        record = {}
        for key in ASIENTOS_REQUIRED_COLUMNS:
            idx = header_index[key]
            record[key] = row[idx] if idx < len(row) else None
        aux_idx = header_index.get('codplanaux')
        record['codplanaux'] = row[aux_idx] if aux_idx is not None and aux_idx < len(row) else None

        record['cod_empresa'] = _normalize_string(record['cod_empresa'])
        record['tipoasiento'] = _normalize_string(record['tipoasiento'])
        record['nrocompr'] = _normalize_string(record['nrocompr'])
        record['nrotransaccion'] = _normalize_string(record['nrotransaccion'])
        record['linea'] = _normalize_string(record['linea'])
        record['codplancta'] = _normalize_string(record['codplancta'])
        record['codplanaux'] = _normalize_string(record.get('codplanaux'))
        record['concepto'] = _normalize_string(record['concepto'])
        record['dbcr'] = _normalize_string(record['dbcr']).upper()
        record['importe'] = _normalize_number(record['importe'])
        record['importeme'] = _normalize_number(record['importeme'])
        record['fecha'] = _normalize_date(record['fecha'])

        row_errors = []
        row_notes = []
        if empresa and record['cod_empresa'] and record['cod_empresa'] != empresa:
            row_errors.append('Cod_Empresa no coincide con la empresa seleccionada.')
        if periodo and record['fecha'] and str(record['fecha'].year) != str(periodo):
            row_errors.append('La fecha no coincide con el periodo seleccionado.')
        if len(record['cod_empresa']) > 2:
            row_errors.append('Cod_Empresa supera 2 caracteres.')
        if not record['tipoasiento']:
            row_errors.append('TipoAsiento vacio.')
        elif len(record['tipoasiento']) > 2:
            row_errors.append('TipoAsiento supera 2 caracteres.')
        elif record['tipoasiento'] not in tipoasiento_map:
            row_errors.append('TipoAsiento no existe.')
        if not record['nrocompr']:
            row_errors.append('NroCompr vacio.')
        elif not _is_integer_like(record['nrocompr']):
            row_errors.append('NroCompr debe ser numerico entero.')
        if not record['nrotransaccion']:
            row_errors.append('NroTransaccion vacio.')
        elif not _is_integer_like(record['nrotransaccion']):
            row_errors.append('NroTransaccion debe ser numerico entero.')
        if not record['linea']:
            row_errors.append('Linea vacia.')
        elif not _is_integer_like(record['linea']):
            row_errors.append('Linea debe ser numerica entera.')
        if not record['codplancta']:
            row_errors.append('CodPlanCta vacio.')
        elif len(record['codplancta']) > 11:
            row_errors.append('CodPlanCta supera 11 caracteres.')
        if record['codplanaux'] and len(record['codplanaux']) > 11:
            row_errors.append('CodPlanAux supera 11 caracteres.')
        if len(record['concepto']) > 120:
            row_errors.append('Concepto supera 120 caracteres.')
        if record['dbcr'] not in ('D', 'C', 'DB', 'CR', 'DEBITO', 'CREDITO'):
            row_errors.append('DbCr invalido.')
        if record['importe'] is None:
            row_errors.append('Importe invalido.')
        if record['importeme'] is None:
            row_errors.append('ImporteME invalido.')
        if record['fecha'] is None:
            row_errors.append('Fecha invalida.')

        if control_row and record['fecha'] is not None:
            fecha_1er_cierre = _normalize_date(control_row.get('Fecha1erCierre', control_row.get('fecha1ercierre', None))) if isinstance(control_row, dict) else None
            fecha_cierre = _normalize_date(control_row.get('FechaCierre', control_row.get('fechacierre', None))) if isinstance(control_row, dict) else None
            mes_1er_cierre = _normalize_int(control_row.get('Mes1erCierre', control_row.get('mes1ercierre', None))) if isinstance(control_row, dict) else None
            mes_cierre = _normalize_int(control_row.get('MesCierre', control_row.get('mescierre', None))) if isinstance(control_row, dict) else None
            mes_fecha = record['fecha'].month
            anho_fecha = record['fecha'].year

            mismo_periodo_cierre = bool(fecha_cierre and fecha_cierre.year == anho_fecha)
            mismo_periodo_1er_cierre = bool(fecha_1er_cierre and fecha_1er_cierre.year == anho_fecha)

            if mismo_periodo_cierre and record['fecha'] <= fecha_cierre:
                row_errors.append('La fecha del asiento cae en un periodo cerrado (FechaCierre: {0}).'.format(fecha_cierre.strftime('%Y-%m-%d')))
            elif mismo_periodo_1er_cierre and record['fecha'] <= fecha_1er_cierre:
                row_errors.append('La fecha del asiento cae en un primer cierre (Fecha1erCierre: {0}).'.format(fecha_1er_cierre.strftime('%Y-%m-%d')))

            if mismo_periodo_cierre and mes_cierre is not None and mes_fecha <= mes_cierre:
                row_errors.append('El mes {0} ya esta cerrado (MesCierre: {1}).'.format(mes_fecha, mes_cierre))
            elif mismo_periodo_1er_cierre and mes_1er_cierre is not None and mes_fecha <= mes_1er_cierre:
                row_errors.append('El mes {0} ya tiene primer cierre (Mes1erCierre: {1}).'.format(mes_fecha, mes_1er_cierre))

        plancta_row = plancta_map.get(record['codplancta'])
        if not plancta_row:
            row_errors.append('CodPlanCta no existe en PLANCTA.')
        else:
            imputable = ''
            if isinstance(plancta_row, dict):
                imputable = _normalize_string(
                    plancta_row.get('imputable', plancta_row.get('Imputable', ''))
                ).upper()

            if imputable == 'S':
                if record['codplanaux']:
                    row_errors.append('La cuenta es imputable y no debe tener auxiliar.')
            else:
                if not record['codplanaux']:
                    row_errors.append('La cuenta requiere auxiliar y CodPlanAux esta vacio.')
                elif (record['codplancta'], record['codplanaux']) not in planaux_map:
                    row_errors.append('CodPlanAux no existe en PLANAUXI para la cuenta.')

        processed_rows.append({
            'fila': rownum,
            'record': record,
            'row_errors': row_errors,
        })

    grouped = {}
    for item in processed_rows:
        record = item['record']
        key = (record['cod_empresa'], record['nrotransaccion'])
        if key not in grouped:
            grouped[key] = {
                'debito': 0.0,
                'credito': 0.0,
                'debitome': 0.0,
                'creditome': 0.0,
                'items': []
            }
        group = grouped[key]
        group['items'].append(item)

        dbcr = record['dbcr']
        importe = float(record['importe'] or 0)
        importeme = float(record['importeme'] or 0)
        if dbcr in ('D', 'DB', 'DEBITO'):
            group['debito'] += importe
            group['debitome'] += importeme
        elif dbcr in ('C', 'CR', 'CREDITO'):
            group['credito'] += importe
            group['creditome'] += importeme

    for key, group in grouped.items():
        local_diff = round(group['debito'] - group['credito'], 4)
        me_diff = round(group['debitome'] - group['creditome'], 4)
        if abs(local_diff) > 0.0001 or abs(me_diff) > 0.0001:
            empresa_key, nrotrans_key = key
            balance_error = (
                'El asiento no balancea para Empresa {0} / NroTransaccion {1}. '
                'Debito={2}, Credito={3}, DebitoME={4}, CreditoME={5}.'
            ).format(
                empresa_key,
                nrotrans_key,
                group['debito'],
                group['credito'],
                group['debitome'],
                group['creditome']
            )
            for item in group['items']:
                item['row_errors'].append(balance_error)

    valid_rows = 0
    for item in processed_rows:
        row_errors = item['row_errors']
        record = item['record']
        rownum = item['fila']

        if row_errors:
            errors.append('Fila {0}: {1}'.format(rownum, ' '.join(row_errors)))
        else:
            valid_rows += 1

        if len(preview_rows) < 100:
            preview_rows.append({
                'fila': rownum,
                'cod_empresa': record['cod_empresa'],
                'tipoasiento': record['tipoasiento'],
                'nrocompr': record['nrocompr'],
                'nrotransaccion': record['nrotransaccion'],
                'linea': record['linea'],
                'codplancta': record['codplancta'],
                'codplanaux': record['codplanaux'],
                'concepto': record['concepto'],
                'dbcr': record['dbcr'],
                'importe': record['importe'],
                'importeme': record['importeme'],
                'fecha': record['fecha'].strftime('%Y-%m-%d') if record['fecha'] else '',
                'estado': 'OK' if not row_errors else 'ERROR',
                'detalle_error': ' | '.join(row_errors),
            })

    return {
        'headers': headers,
        'preview_rows': preview_rows,
        'errors': errors[:200],
        'cleaned_rows': [
            {
                'fila': item['fila'],
                'cod_empresa': item['record']['cod_empresa'],
                'tipoasiento': item['record']['tipoasiento'],
                'nrocompr': item['record']['nrocompr'],
                'nrotransaccion': item['record']['nrotransaccion'],
                'linea': item['record']['linea'],
                'codplancta': item['record']['codplancta'],
                'codplanaux': item['record']['codplanaux'],
                'concepto': item['record']['concepto'],
                'dbcr': item['record']['dbcr'],
                'importe': item['record']['importe'],
                'importeme': item['record']['importeme'],
                'fecha': item['record']['fecha'].strftime('%Y-%m-%d') if item['record']['fecha'] else '',
                'row_errors': item['row_errors'],
            }
            for item in processed_rows
        ],
        'summary': {
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': max(0, total_rows - valid_rows),
        }
    }


def _validate_ventas_rows(rows, empresa, periodo, tipo_comprobante='', sucursal=''):
    selected_errors = []
    if not empresa:
        selected_errors.append('Debe seleccionar una empresa.')
    if not periodo:
        selected_errors.append('Debe seleccionar un periodo.')
    if not tipo_comprobante:
        selected_errors.append('Debe seleccionar un tipo de comprobante.')
    if not sucursal:
        selected_errors.append('Debe seleccionar una sucursal.')
    if selected_errors:
        return {
            'headers': [],
            'preview_rows': [],
            'errors': selected_errors,
            'summary': {'total_rows': 0, 'valid_rows': 0, 'error_rows': 0},
            'cleaned_rows': [],
        }

    if not rows:
        return {
            'headers': [],
            'preview_rows': [],
            'errors': ['El archivo no contiene filas.'],
            'summary': {'total_rows': 0, 'valid_rows': 0, 'error_rows': 0},
            'cleaned_rows': [],
        }

    headers = [_normalize_excel_header(x, VENTAS_HEADER_ALIASES) for x in rows[0]]
    missing = [col for col in VENTAS_REQUIRED_COLUMNS if col not in headers]
    if missing:
        return {
            'headers': headers,
            'preview_rows': [],
            'errors': ['Faltan columnas obligatorias: ' + ', '.join(missing)],
            'summary': {'total_rows': 0, 'valid_rows': 0, 'error_rows': 0},
            'cleaned_rows': [],
        }

    header_index = {name: idx for idx, name in enumerate(headers)}
    preview_rows = []
    cleaned_rows = []
    errors = []
    total_rows = 0
    valid_rows = 0
    references = _load_ventas_reference_data(empresa, periodo, rows, tipo_comprobante, sucursal)
    plancta_map = references.get('plancta', {})
    planaux_map = references.get('planauxi', {})
    articulo_set = references.get('articulos', set())
    condiciones = references.get('condiciones', set())
    monedas = references.get('monedas', set())
    clientes_map = references.get('clientes', {})
    created_clientes = references.get('created_clientes', set())
    missing_clientes = references.get('missing_clientes', set())
    control_row = references.get('control')

    for excel_row_num, row in enumerate(rows[1:], start=2):
        if not any(str(v or '').strip() for v in row):
            continue
        total_rows += 1

        record = {}
        row_errors = []
        row_notes = []
        for key in VENTAS_REQUIRED_COLUMNS:
            idx = header_index[key]
            record[key] = row[idx] if idx < len(row) else None

        record['fha_cbte'] = _normalize_date(record['fha_cbte'])
        record['linea'] = _normalize_int(record['linea'])
        record['comp_numero'] = _normalize_string(record['comp_numero'])
        record['razon_social'] = _normalize_string(record['razon_social'])
        record['ruc'] = _normalize_ruc(record['ruc'])
        record['cta_cliente'] = _normalize_string(record['cta_cliente'])
        record['aux_cliente'] = _normalize_string(record['aux_cliente'])
        record['cod_articulo'] = _normalize_string(record['cod_articulo'])
        record['descrip'] = _normalize_string(record['descrip'])
        record['cantidad'] = _normalize_number(record['cantidad'])
        record['pr_unit'] = _normalize_number(record['pr_unit'])
        record['total'] = _normalize_number(record['total'])
        record['cod_con_vta'] = _normalize_string(record['cod_con_vta'])
        record['codmoneda'] = _normalize_string(record['codmoneda'])
        record['fact_cambio'] = _normalize_number(record['fact_cambio'])
        record['direccion'] = _normalize_string(record['direccion'])
        record['telefono'] = _normalize_string(record['telefono'])
        record['observ'] = _normalize_string(record['observ'])
        record['to_gravado_5'] = _normalize_number(record['to_gravado_5'])
        record['total_iva_5'] = _normalize_number(record['total_iva_5'])
        record['to_gravado_10'] = _normalize_number(record['to_gravado_10'])
        record['total_iva_10'] = _normalize_number(record['total_iva_10'])
        record['exenta'] = _normalize_number(record['exenta'])
        record['anulado'] = _normalize_string(record['anulado']).upper() or 'N'
        record['anulado_el'] = _normalize_string(record['anulado_el'])
        record['anulado_por'] = _normalize_string(record['anulado_por'])
        record['total_neto'] = _normalize_number(record['total_neto'])
        record['porcpartgravado'] = _normalize_number(record['porcpartgravado'])
        record['montopartgravado'] = _normalize_number(record['montopartgravado'])
        record['nrotimb'] = _normalize_number(record['nrotimb'])
        record['cod_establecimiento'] = _normalize_string(record['cod_establecimiento'])
        record['cod_ptoexpedicion'] = _normalize_string(record['cod_ptoexpedicion'])
        record['cuotas'] = _normalize_int(record['cuotas'])
        record['cod_empresa'] = empresa
        record['periodo'] = periodo
        record['cod_tp_comp'] = tipo_comprobante
        record['cod_sucursal'] = sucursal

        if not record['razon_social']:
            record['razon_social'] = 'Sin Nombre'
        if not record['ruc']:
            record['ruc'] = '44444401-7'
        record['cod_cliente'] = clientes_map.get(record['ruc'], '')

        if record['fha_cbte'] is None:
            row_errors.append('Fecha comprobante invalida.')
        if record['linea'] is None:
            row_errors.append('Linea invalida.')
        if not record['comp_numero']:
            row_errors.append('Comp_Numero es obligatorio.')
        if not record['cta_cliente']:
            row_errors.append('Cta_Cliente es obligatoria.')
        elif record['cta_cliente'] not in plancta_map:
            row_errors.append('Cta_Cliente no existe en PlanCta.')
        if not record['cod_articulo']:
            row_errors.append('Cod_Articulo es obligatorio.')
        elif articulo_set and record['cod_articulo'] not in articulo_set:
            row_errors.append('El articulo con el codigo {0} no existe favor cargar.'.format(record['cod_articulo']))
        if record['ruc'] in missing_clientes and record['ruc'] not in created_clientes:
            row_notes.append('Cliente no existe, se creara automaticamente.')
        elif record['ruc'] in missing_clientes:
            row_notes.append('Cliente no existe, se creara automaticamente.')
        elif record['ruc'] in created_clientes:
            row_notes.append('Cliente creado automaticamente.')
        if record['cantidad'] is None:
            row_errors.append('Cantidad invalida.')
        elif record['cantidad'] == 0:
            row_errors.append('Cantidad no puede ser 0.')
        if record['pr_unit'] is None:
            row_errors.append('Pr_Unit invalido.')
        elif record['pr_unit'] == 0:
            row_errors.append('Pr_Unit no puede ser 0.')
        if record['total'] is None:
            row_errors.append('Total invalido.')
        elif record['total'] == 0:
            row_errors.append('Total no puede ser 0.')
        if record['cantidad'] and record['total'] == 0:
            row_errors.append('Total no puede ser 0 si Cantidad es mayor a 0.')
        if record['pr_unit'] and record['total'] == 0:
            row_errors.append('Total no puede ser 0 si Pr_Unit es mayor a 0.')
        if record['cantidad'] and record['pr_unit'] == 0:
            row_errors.append('Pr_Unit no puede ser 0 si Cantidad es mayor a 0.')
        if record['cod_con_vta'] == '':
            row_errors.append('Cod_Con_Vta es obligatorio.')
        elif condiciones and record['cod_con_vta'] not in condiciones:
            row_errors.append('Cod_Con_Vta no existe en Terminos.')
        if not record['codmoneda']:
            row_errors.append('CodMoneda es obligatorio.')
        elif monedas and record['codmoneda'].upper() not in monedas:
            row_errors.append('CodMoneda no existe.')
        if record['fact_cambio'] is None:
            row_errors.append('Fact_Cambio invalido.')
        if record['total_neto'] is None:
            row_errors.append('Total_Neto invalido.')
        if record['to_gravado_5'] is None:
            row_errors.append('To_Gravado_5 invalido.')
        if record['total_iva_5'] is None:
            row_errors.append('Total_Iva_5 invalido.')
        if record['to_gravado_10'] is None:
            row_errors.append('To_Gravado_10 invalido.')
        if record['total_iva_10'] is None:
            row_errors.append('Total_Iva_10 invalido.')
        if record['exenta'] is None:
            row_errors.append('Exenta invalida.')
        if record['porcpartgravado'] is None:
            row_errors.append('PorcPartGravado invalido.')
        if record['montopartgravado'] is None:
            row_errors.append('MontoPartGravado invalido.')
        if record['nrotimb'] is None:
            row_errors.append('NroTimb invalido.')
        if not record['cod_establecimiento']:
            row_errors.append('Cod_Establecimiento es obligatorio.')
        if not record['cod_ptoexpedicion']:
            row_errors.append('Cod_PtoExpedicion es obligatorio.')
        if record['cuotas'] is None:
            row_errors.append('Cuotas invalidas.')
        if record['anulado'] not in ('S', 'N'):
            row_errors.append('Anulado debe ser S o N.')

        if record['fha_cbte'] and control_row:
            cierre_fecha = control_row.get('FechaCierre', control_row.get('fechacierre',
                             control_row.get('FECHACIERRE', control_row.get('fecha_cierre'))))
            cierre_mes = control_row.get('MesCierre', control_row.get('mescierre',
                           control_row.get('MESCierre', control_row.get('MESCIERRE', control_row.get('mes_cierre')))))
            primer_cierre_fecha = control_row.get('Fecha1erCierre', control_row.get('fecha1ercierre',
                                   control_row.get('FECHA1ERCIERRE', control_row.get('fecha_1er_cierre'))))
            primer_cierre_mes = control_row.get('Mes1erCierre', control_row.get('mes1ercierre',
                                 control_row.get('MES1ERCIERRE', control_row.get('mes_1er_cierre'))))
            fecha_excel = record['fha_cbte']
            if isinstance(fecha_excel, str):
                fecha_excel = _normalize_date(fecha_excel)

            def _same_period(dt):
                try:
                    return str(dt.year) == str(periodo)
                except Exception:
                    return False

            def _is_closed(fecha, mes):
                if fecha and fecha_excel and _same_period(fecha) and fecha_excel <= fecha:
                    return True
                if mes and fecha_excel and _same_period(fecha_excel):
                    try:
                        return fecha_excel.month <= int(mes)
                    except Exception:
                        return False
                return False

            cierre_fecha_dt = _normalize_date(cierre_fecha)
            primer_cierre_fecha_dt = _normalize_date(primer_cierre_fecha)
            cierre_fecha_has_value = bool(_normalize_string(cierre_fecha))

            cierre_es_periodo_actual = False
            if cierre_fecha_dt and _same_period(cierre_fecha_dt):
                cierre_es_periodo_actual = True
            elif cierre_mes and fecha_excel and _same_period(fecha_excel) and not cierre_fecha_has_value:
                cierre_es_periodo_actual = True

            if cierre_es_periodo_actual and _is_closed(cierre_fecha_dt, cierre_mes):
                msg = 'La fecha del comprobante cae en un periodo cerrado'
                if cierre_fecha_dt:
                    msg += ' (FechaCierre: {0})'.format(cierre_fecha_dt.strftime('%d/%m/%Y'))
                elif cierre_mes:
                    msg += ' (MesCierre: {0})'.format(cierre_mes)
                row_errors.append(msg + '.')
            elif _is_closed(primer_cierre_fecha_dt, primer_cierre_mes):
                msg = 'La fecha del comprobante cae en un primer cierre'
                if primer_cierre_fecha_dt:
                    msg += ' (Fecha1erCierre: {0})'.format(primer_cierre_fecha_dt.strftime('%d/%m/%Y'))
                elif primer_cierre_mes:
                    msg += ' (Mes1erCierre: {0})'.format(primer_cierre_mes)
                row_errors.append(msg + '.')

        plancta_row = plancta_map.get(record['cta_cliente'])
        if plancta_row:
            imputable = _normalize_string(plancta_row.get('Imputable', plancta_row.get('imputable', ''))).upper()
            auxiliar = _normalize_string(plancta_row.get('Auxiliar', plancta_row.get('auxiliar', ''))).upper()
            if imputable == 'S':
                if record['aux_cliente']:
                    row_errors.append('La cuenta es imputable y no debe tener Aux_Cliente.')
            else:
                if not record['aux_cliente']:
                    row_errors.append('La cuenta requiere Aux_Cliente y esta vacio.')
                elif (record['cta_cliente'], record['aux_cliente']) not in planaux_map:
                    row_errors.append('Aux_Cliente no existe en PlanAuxi para la cuenta.')

        if row_errors:
            errors.append('Fila {0}: {1}'.format(excel_row_num, ' '.join(row_errors)))
        else:
            valid_rows += 1
            cleaned = dict(record)
            if cleaned.get('fha_cbte'):
                try:
                    cleaned['fha_cbte'] = cleaned['fha_cbte'].strftime('%Y-%m-%d')
                except Exception:
                    cleaned['fha_cbte'] = str(cleaned['fha_cbte'])
            cleaned_rows.append(cleaned)

        if len(preview_rows) < 100:
            preview_rows.append({
                'fila': excel_row_num,
                'cod_empresa': record['cod_empresa'],
                'fha_cbte': record['fha_cbte'].strftime('%Y-%m-%d') if record['fha_cbte'] else '',
                'linea': record['linea'],
                'comp_numero': record['comp_numero'],
                'razon_social': record['razon_social'],
                'ruc': record['ruc'],
                'cta_cliente': record['cta_cliente'],
                'aux_cliente': record['aux_cliente'],
                'cod_articulo': record['cod_articulo'],
                'descrip': record['descrip'],
                'cantidad': record['cantidad'],
                'pr_unit': record['pr_unit'],
                'total': record['total'],
                'cod_con_vta': record['cod_con_vta'],
                'codmoneda': record['codmoneda'],
                'fact_cambio': record['fact_cambio'],
                'direccion': record['direccion'],
                'telefono': record['telefono'],
                'observ': record['observ'],
                'to_gravado_5': record['to_gravado_5'],
                'total_iva_5': record['total_iva_5'],
                'to_gravado_10': record['to_gravado_10'],
                'total_iva_10': record['total_iva_10'],
                'exenta': record['exenta'],
                'anulado': record['anulado'],
                'anulado_el': record['anulado_el'],
                'anulado_por': record['anulado_por'],
                'total_neto': record['total_neto'],
                'porcpartgravado': record['porcpartgravado'],
                'montopartgravado': record['montopartgravado'],
                'nrotimb': record['nrotimb'],
                'cod_establecimiento': record['cod_establecimiento'],
                'cod_ptoexpedicion': record['cod_ptoexpedicion'],
                'cuotas': record['cuotas'],
                'cod_tp_comp': record['cod_tp_comp'],
                'cod_sucursal': record['cod_sucursal'],
                'estado': 'OK' if not row_errors else 'ERROR',
                'detalle_error': ' | '.join(row_errors) if row_errors else ' | '.join(row_notes),
            })

    return {
        'headers': headers,
        'preview_rows': preview_rows,
        'errors': errors[:200],
        'cleaned_rows': cleaned_rows,
        'summary': {
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'error_rows': max(0, total_rows - valid_rows),
        },
        'created_clientes_count': len(created_clientes),
    }


def server_error(request, template_name='500.html'):
    return render(request, template_name, status=500)


def custom_404(request, exception=None):
    return render(request, '404.html', status=404)

def custom_403(request, exception=None):
    return render(request, '403.html', status=403)


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = auth.authenticate(request, username=username, password=password)
        if user is not None:
            auth.login(request, user)
            # Close any previous active sessions for this user to avoid duplicates
            current_key = request.session.session_key
            for s in Session.objects.filter(expire_date__gt=timezone.now()):
                data = s.get_decoded()
                if str(data.get('_auth_user_id')) == str(user.id):
                    if current_key and s.session_key == current_key:
                        continue
                    s.delete()
            return redirect('/accounts/loggedin/')
        else:
            return redirect('/accounts/invalid/')
    return render(request, 'login.html')


@login_required
def loggedin(request):
    modulos_usuario = []
    permisos = [[p[1], p[0]] for p in custom_permissions.models.permissions_list]

    iconos = {
        "Finanzas": "fa fa-university fa-5x",
        "Ventas": "fa fa-file-text-o fa-5x",
        "Stock": "fa fa-cube fa-5x",
        "Compras": "fa fa-shopping-cart fa-5x",
        "RRHH": "fa fa-users fa-5x",
        "Admin": "fa fa-cogs fa-5x",
        "Migraciones": "fa fa-random fa-5x",
        "Produccion": "fa fa-industry fa-5x",
    }

    for permiso in permisos:
        if request.user.has_perm('custom_permissions.' + permiso[1]):
            icono = iconos.get(permiso[0], '')
            modulos_usuario.append([permiso[0], icono])

    return render(request, 'loggedin.html', {'modulos_usuario': modulos_usuario})


@login_required
def menu(request):
    q = request.GET.get('q', '')
    request.session['modulo_seleccionado'] = q
    permisos_modulos = {p[1]: p[0] for p in custom_permissions.models.permissions_list}
    permisos_informes = {p[0]: p[1] for p in custom_permissions.models.permissions_informes}

    titulo = ''
    items = []
    grupos = []
    if request.user.has_perm('custom_permissions.' + permisos_modulos.get(q, '')):
        titulo = q

        menu_map = {
            'Finanzas': [
                {'perm': 'informe_mayorcuenta', 'url': '/finanzas/mayor_cuentas/'},
                {'perm': 'informe_mayorcuentaauxiliar', 'url': '/finanzas/mayor_cuentas_auxiliar/'},
                {'perm': 'informe_ordenpago', 'url': '/finanzas/ordenpago/'},
                {'perm': 'informe_centrocostos', 'url': '/finanzas/centro_costos/'},
                {'perm': 'informe_extractocuenta', 'url': '/finanzas/extracto_cuentas_banco/'},
                {'perm': 'informe_flujofondo', 'url': '/finanzas/flujo_fondo/'},
                {'perm': 'informe_balancegeneral', 'url': '/finanzas/balancegeneral/'},
                {'perm': 'informe_balancegeneralpuc', 'url': '/finanzas/balancegeneral_puc/'},
                {'perm': 'informe_balancegeneralcomprobado', 'url': '/finanzas/balancegeneral_comprobado/'},
                {'perm': 'informe_diario', 'url': '/finanzas/diario_comprobado/'},
                {'perm': 'informe_bienactivo', 'url': '/finanzas/bienes_activo/'},
                {'perm': None, 'titulo': 'Dashboard financiero', 'url': '/finanzas/dashboard_financiero/'},
                {'perm': None, 'titulo': 'Balance Integral', 'url': '/finanzas/balance_integral/'},
            ],
            'Ventas': [
                {'perm': 'informe_ventascomprobante', 'url': '/ventas/resumido/'},
                {'perm': 'informe_presupuesto', 'url': '/ventas/presupuesto/'},
                {'perm': 'informe_cuentascobrar', 'url': '/ventas/cuentas_cobrar/'},
                {'perm': 'informe_recaudaciones', 'url': '/ventas/recaudaciones/'},
                {'perm': 'informe_ventaestadisticos', 'url': '/ventas/estadisticas/'},
            ],
            'Compras': [
                {'perm': 'informe_ordencompra', 'url': '/compras/ordencompra/'},
                {'perm': 'informe_compras', 'url': '/compras/compras/'},
                {'perm': 'informe_comprasestadisticos', 'url': '/compras/estadisticos/'},
                {'perm': 'informe_saldosproveedores', 'url': '/compras/saldos_proveedores/'},
                {'perm': 'informe_fondofijo', 'url': '/compras/fondo_fijo/'},
                {'perm': 'informe_gastosrendir', 'url': '/compras/gastos_rendir/'},
            ],
            'Stock': [
                {'perm': 'informe_articulos', 'url': '/stock/articulos/'},
                {'perm': 'informe_precios', 'url': '/stock/precios/'},
                {'perm': 'informe_depositos', 'url': '/stock/existencia_deposito/'},
                {'perm': 'informe_valorizado', 'url': '/stock/valorizado/'},
                {'perm': None, 'titulo': 'Costo Articulo Full', 'url': '/stock/costo_articulo_full/'},
            ],
            'RRHH': [
                {'perm': 'informe_legajos', 'url': '/rrhh/legajos/'},
                {'perm': 'informe_sueldosjornales', 'url': '/rrhh/sueldos_jornales/'},
                {'perm': 'informe_anticipos', 'url': '/rrhh/anticipos/'},
                {'perm': 'informe_aguinaldos', 'url': '/rrhh/aguinaldos/'},
                {'perm': 'informe_recibos', 'url': '/rrhh/recibos/'},
            ],
            'Admin': [
                {'perm': None, 'titulo': 'Usuarios', 'url': '/admin/auth/user/'},
                {'perm': None, 'titulo': 'Grupos', 'url': '/admin/auth/group/'},
                {'perm': None, 'titulo': 'Asignación de Empresas', 'url': '/custom_permissions/asignar_empresa_usuario/'},
                {'perm': None, 'titulo': 'Usuarios Conectados', 'url': '/custom_permissions/active_sessions/'},
                {'perm': None, 'titulo': 'Configuración de Base de Datos', 'url': '/custom_permissions/db_config/'},
                {'perm': None, 'titulo': 'Configuración de Email', 'url': '/custom_permissions/email_config/'},
            ],
            'Migraciones': [
                {'perm': None, 'titulo': 'Asientos', 'url': '/migraciones/asientos/'},
                {'perm': None, 'titulo': 'Ventas', 'url': '/migraciones/ventas/'},
                {'perm': None, 'titulo': 'Compras', 'url': '/migraciones/compras/'},
            ],
        }

        for item in menu_map.get(q, []):
            perm = item.get('perm')
            if perm:
                if request.user.has_perm('custom_permissions.' + perm):
                    items.append({
                        'titulo': item.get('titulo') or permisos_informes.get(perm, perm),
                        'url': item.get('url', '#'),
                        'perm': perm,
                    })
            else:
                items.append({
                    'titulo': item.get('titulo', 'Item'),
                    'url': item.get('url', '#'),
                    'perm': None,
                })

        if items:
            seen = set()
            unique_items = []
            for item in items:
                key = (item.get('url'), item.get('titulo'))
                if key in seen:
                    continue
                seen.add(key)
                unique_items.append(item)
            items = unique_items

        if q == 'Finanzas':
            grupos_def = [
                ('Balances', [
                    'informe_balancegeneral',
                    'informe_balancegeneralpuc',
                    'informe_balancegeneralcomprobado',
                ]),
                ('Libros', [
                    'informe_mayorcuenta',
                    'informe_mayorcuentaauxiliar',
                    'informe_diario',
                ]),
                ('Pagos', [
                    'informe_ordenpago',
                ]),
                ('Otros', [
                    'informe_centrocostos',
                    'informe_extractocuenta',
                    'informe_flujofondo',
                    'informe_bienactivo',
                ]),
            ]

            items_by_perm = {i['perm']: i for i in items}
            for nombre, perms in grupos_def:
                hijos = []
                for perm in perms:
                    item = items_by_perm.get(perm)
                    if item:
                        hijos.append(item)
                if nombre == 'Balances':
                    for item in items:
                        if item.get('url') == '/finanzas/balance_integral/':
                            hijos.append(item)
                            break
                if nombre == 'Otros':
                    for item in items:
                        if item.get('url') == '/finanzas/dashboard_financiero/':
                            hijos.append(item)
                            break
                if hijos:
                    grupos.append({'titulo': nombre, 'items': hijos})

    return render(request, 'menu.html', {
        'titulo': titulo,
        'items': items,
        'grupos': grupos,
        'modulo': q,
    })


@login_required
def migraciones_asientos(request):
    session_key = 'migraciones_asientos_validated'
    if request.method == 'GET':
        request.session.pop(session_key, None)
    saved = request.session.get(session_key) or {}

    empresas_codigos = list(
        custom_permissions.models.UsuarioEmpresa.objects
        .filter(user=request.user, db='Integrado')
        .order_by('empresa')
        .values_list('empresa', flat=True)
        .distinct()
    )
    empresas = [{'codigo': e, 'nombre': e} for e in empresas_codigos]
    selected_empresa = request.POST.get('empresa') or request.GET.get('empresa') or ''
    selected_periodo = request.POST.get('periodo') or request.GET.get('periodo') or str(timezone.now().year)
    upload_result = None
    upload_error = ''
    validation_ok = False
    validation_message = ''
    migration_ok = False
    migration_message = ''
    migration_process = []
    action = request.POST.get('action', '')
    if saved.get('rows'):
        validation_ok = True
        validation_message = 'Excel validado. Puede migrar o limpiar la pantalla.'
        upload_result = {
            'summary': {
                'total_rows': len(saved.get('rows', [])),
                'valid_rows': len(saved.get('rows', [])),
                'error_rows': 0,
            },
            'preview_rows': [],
            'errors': [],
        }

    if request.method == 'POST' and action == 'validar' and request.FILES.get('archivo_excel'):
        uploaded_file = request.FILES['archivo_excel']
        try:
            rows = _read_asientos_excel(uploaded_file)
            upload_result = _validate_asientos_rows(rows, selected_empresa, selected_periodo)
            summary = upload_result.get('summary', {}) if upload_result else {}
            if summary.get('total_rows', 0) > 0 and summary.get('error_rows', 0) == 0:
                validation_ok = True
                validation_message = 'Excel validado por Django. No se detectaron errores en el control previo.'
                request.session[session_key] = {
                    'empresa': selected_empresa,
                    'periodo': selected_periodo,
                    'rows': upload_result.get('cleaned_rows', []),
                }
                request.session.modified = True
            elif summary.get('total_rows', 0) > 0:
                validation_message = 'Excel controlado por Django con observaciones. Revise los errores antes de migrar.'
                request.session.pop(session_key, None)
        except Exception as exc:
            upload_error = str(exc)
            validation_message = 'No se pudo validar el archivo.'
            request.session.pop(session_key, None)

    elif request.method == 'POST' and action == 'migrar':
        if not saved:
            upload_error = 'No hay un Excel validado listo para migrar. Primero valide el archivo.'
        elif saved.get('empresa') != selected_empresa or saved.get('periodo') != selected_periodo:
            upload_error = 'La empresa o el periodo ya no coinciden con la validacion guardada. Limpie la pantalla y valide nuevamente.'
            request.session.pop(session_key, None)
        else:
            try:
                saved_rows = saved.get('rows', [])
                batches = _chunk_asientos_rows(saved_rows)
                total_batches = len(batches)
                total_inserted = 0
                inserted_ranges = []
                first_inserted_nro = None
                last_inserted_nro = None
                migration_process = [
                    'Migracion preparada en {0} lote(s) de hasta {1} asiento(s) por lote.'.format(
                        total_batches,
                        ASIENTOS_MIGRATION_BATCH_SIZE
                    )
                ]

                for index, batch_rows in enumerate(batches, start=1):
                    first_row = batch_rows[0] if batch_rows else {}
                    lote_empresa = first_row.get('cod_empresa') or selected_empresa
                    lote_nro = first_row.get('nrotransaccion') or ''
                    migration_process.append(
                        'Enviando lote {0} de {1} para Empresa {2} / NroTransaccion {3} con {4} fila(s).'.format(
                            index, total_batches, lote_empresa, lote_nro, len(batch_rows)
                        )
                    )
                    try:
                        resp = _node_post(
                            'migraciones/asientos/import',
                            json={
                                'empresa': selected_empresa,
                                'periodo': selected_periodo,
                                'rows': batch_rows,
                                'user_id': request.user.id,
                                'username': request.user.username,
                            },
                            timeout=NODE_API_TIMEOUT_LONG
                        )
                    except Exception as exc:
                        migration_process.append(
                            'Fallo el lote {0} de {1}: {2}'.format(index, total_batches, str(exc))
                        )
                        upload_error = (
                            'La migracion se interrumpio en el lote {0} de {1}. '
                            'Los lotes anteriores ya pudieron haberse insertado.'
                        ).format(index, total_batches)
                        break

                    data = resp.json() if resp.content else {}
                    batch_process = data.get('process') or []
                    for step in batch_process:
                        migration_process.append('[Lote {0}] {1}'.format(index, step))

                    if resp.status_code >= 400 or not data.get('ok'):
                        migration_process.append(
                            'El lote {0} devolvio error y se detuvo la migracion.'.format(index)
                        )
                        upload_error = data.get('error') or (
                            'No se pudo migrar el lote {0} de {1}.'.format(index, total_batches)
                        )
                        break

                    total_inserted += int(data.get('inserted_count') or 0)
                    if data.get('from_nrotransac') and data.get('to_nrotransac'):
                        inserted_ranges.append(
                            '{0}-{1}'.format(data.get('from_nrotransac'), data.get('to_nrotransac'))
                        )
                        if first_inserted_nro is None:
                            first_inserted_nro = data.get('from_nrotransac')
                        last_inserted_nro = data.get('to_nrotransac')
                    migration_process.append(
                        'Lote {0} completado correctamente.'.format(index)
                    )

                if not upload_error:
                    migration_ok = True
                    migration_message = (
                        'Migracion finalizada correctamente. '
                        'Lotes procesados: {0}. Asientos insertados: {1}.'
                    ).format(total_batches, total_inserted)
                    if first_inserted_nro is not None and last_inserted_nro is not None:
                        migration_message += ' Desde Asiento Nro.: {0} - Hasta Nro.: {1}.'.format(
                            first_inserted_nro,
                            last_inserted_nro
                        )
                    request.session.pop(session_key, None)
            except Exception as exc:
                upload_error = str(exc)
                migration_process = [
                    'Preparando migracion para empresa {0} y periodo {1}.'.format(selected_empresa, selected_periodo),
                    'Enviando datos validados al servicio de migracion.',
                    'La conexion con el servicio fue interrumpida antes de responder.',
                    'Revise el proceso del servidor Node o vuelva a intentar la migracion.',
                ]

    return render(request, 'migraciones/asientos.html', {
        'titulo': 'Migraciones',
        'subtitulo': 'Asientos',
        'user': request.user.id,
        'empresas': empresas,
        'selected_empresa': selected_empresa,
        'selected_periodo': selected_periodo,
        'upload_result': upload_result,
        'upload_error': upload_error,
        'validation_ok': validation_ok,
        'validation_message': validation_message,
        'migration_ok': migration_ok,
        'migration_message': migration_message,
        'migration_process': migration_process,
    })


@login_required
def migraciones_ventas(request):
    session_key = 'migraciones_ventas_validated'
    if request.method == 'GET':
        request.session.pop(session_key, None)

    selected_empresa = request.POST.get('empresa') or request.GET.get('empresa') or ''
    selected_periodo = request.POST.get('periodo') or request.GET.get('periodo') or str(timezone.now().year)
    selected_empresa_desc = request.POST.get('empresa_desc') or request.GET.get('empresa_desc') or ''
    selected_tipo = request.POST.get('tipo_comprobante') or request.GET.get('tipo_comprobante') or ''
    selected_sucursal = request.POST.get('sucursal') or request.GET.get('sucursal') or ''
    upload_result = None
    upload_error = ''
    validation_ok = False
    migration_ok = False
    migration_message = ''
    migration_process = []
    importacion_habilitada = True
    importacion_mensaje = ''
    action = request.POST.get('action', '')

    if selected_empresa:
        try:
            meta_resp = _node_get('empresas/{0}/importar_ventas'.format(selected_empresa))
            meta_resp.raise_for_status()
            meta_data = meta_resp.json().get('data', [])
            row = meta_data[0] if meta_data else {}
            importar_ventas = _normalize_string(
                row.get('importar_ventas', row.get('IMPORTAR_VENTAS', 'N'))
            ).upper()
            if importar_ventas != 'S':
                importacion_habilitada = False
                importacion_mensaje = 'La empresa seleccionada no tiene habilitada la opcion Importar Ventas.'
        except Exception:
            importacion_habilitada = False
            importacion_mensaje = 'No se pudo verificar si la empresa tiene habilitada la opcion Importar Ventas.'

    if request.method == 'POST' and action == 'validar' and request.FILES.get('archivo_excel'):
        uploaded_file = request.FILES['archivo_excel']
        if not importacion_habilitada:
            upload_error = importacion_mensaje
        else:
            try:
                rows = _read_asientos_excel(uploaded_file)
                upload_result = _validate_ventas_rows(
                    rows,
                    selected_empresa,
                    selected_periodo,
                    selected_tipo,
                    selected_sucursal,
                )
                summary = upload_result.get('summary', {}) if upload_result else {}
                if summary.get('total_rows', 0) > 0 and summary.get('error_rows', 0) == 0:
                    validation_ok = True
                    request.session[session_key] = {
                        'rows': upload_result.get('cleaned_rows', []),
                        'empresa': selected_empresa,
                        'periodo': selected_periodo,
                        'tipo': selected_tipo,
                        'sucursal': selected_sucursal,
                        'created_clientes_count': upload_result.get('created_clientes_count', 0),
                        'empresa_desc': selected_empresa_desc,
                    }
                    request.session.modified = True
            except Exception as exc:
                upload_error = str(exc)

    elif request.method == 'POST' and action == 'migrar':
        payload = request.session.get(session_key) or {}
        rows = payload.get('rows', [])
        if not rows:
            upload_error = 'No hay datos validados para migrar.'
        else:
            try:
                resp = _node_post(
                    'migraciones/ventas/import',
                    json={
                        'empresa': payload.get('empresa', selected_empresa),
                        'empresa_desc': payload.get('empresa_desc', selected_empresa_desc),
                        'periodo': payload.get('periodo', selected_periodo),
                        'cod_tp_comp': payload.get('tipo', selected_tipo),
                        'cod_sucursal': payload.get('sucursal', selected_sucursal),
                        'rows': rows,
                        'username': request.user.username,
                    },
                    timeout=NODE_API_TIMEOUT_LONG
                )
                resp.raise_for_status()
                data = resp.json() or {}
                migration_ok = bool(data.get('ok'))
                if migration_ok:
                    migration_message = 'Ventas grabadas correctamente.'
                else:
                    migration_message = data.get('message') or data.get('error', '')
                migration_process = data.get('process', []) or []
                created_count = payload.get('created_clientes_count', 0) or 0
                if created_count:
                    migration_process = ['Clientes creados automaticamente: {0}.'.format(created_count)] + migration_process
                if migration_ok:
                    request.session.pop(session_key, None)
            except Exception as exc:
                upload_error = str(exc)
                migration_process = [
                    'Preparando migracion para empresa {0} y periodo {1}.'.format(selected_empresa, selected_periodo),
                    'Enviando datos validados al servicio de migracion.',
                    'La conexion con el servicio fue interrumpida antes de responder.',
                    'Revise el proceso del servidor Node o vuelva a intentar la migracion.',
                ]

    return render(request, 'migraciones/ventas.html', {
        'titulo': 'Migraciones',
        'subtitulo': 'Ventas',
        'user': request.user.id,
        'selected_empresa': selected_empresa,
        'selected_periodo': selected_periodo,
        'selected_tipo': selected_tipo,
        'selected_sucursal': selected_sucursal,
        'upload_result': upload_result,
        'upload_error': upload_error,
        'validation_ok': validation_ok,
        'importacion_habilitada': importacion_habilitada,
        'importacion_mensaje': importacion_mensaje,
        'migration_ok': migration_ok,
        'migration_message': migration_message,
        'migration_process': migration_process,
    })


@login_required
def migraciones_compras(request):
    return render(request, 'migraciones/simple.html', {
        'titulo': 'Migraciones',
        'subtitulo': 'Compras',
    })


def invalid_login(request):
    return render(request, 'invalid_login.html')


def logout_view(request):
    auth.logout(request)
    return render(request, 'logout.html')


def _json_http(payload, status=200):
    return HttpResponse(
        json.dumps(payload, default=str),
        status=status,
        content_type='application/json'
    )


@csrf_exempt
@require_POST
def api_migraciones_asientos_validate(request):
    selected_empresa = request.POST.get('empresa') or ''
    selected_periodo = request.POST.get('periodo') or str(timezone.now().year)
    uploaded_file = request.FILES.get('archivo_excel')

    if not selected_empresa:
        return _json_http({'ok': False, 'message': 'Seleccione una empresa.'}, status=400)

    if not selected_periodo:
        return _json_http({'ok': False, 'message': 'Seleccione un periodo.'}, status=400)

    if not uploaded_file:
        return _json_http({'ok': False, 'message': 'Adjunte un archivo Excel.'}, status=400)

    try:
        rows = _read_asientos_excel(uploaded_file)
        upload_result = _validate_asientos_rows(rows, selected_empresa, selected_periodo)
        summary = upload_result.get('summary', {}) if upload_result else {}
        validation_ok = summary.get('total_rows', 0) > 0 and summary.get('error_rows', 0) == 0

        return _json_http({
            'ok': True,
            'validation_ok': validation_ok,
            'message': (
                'Excel validado correctamente.'
                if validation_ok
                else 'Excel controlado con observaciones. Revise los errores antes de migrar.'
            ),
            'data': upload_result,
        })
    except Exception as exc:
        return _json_http({'ok': False, 'message': str(exc)}, status=400)


@csrf_exempt
@require_POST
def api_migraciones_asientos_migrate(request):
    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return _json_http({'ok': False, 'message': 'No se pudo leer la solicitud.'}, status=400)

    selected_empresa = body.get('empresa') or ''
    selected_periodo = body.get('periodo') or str(timezone.now().year)
    rows = body.get('rows') or []
    username = body.get('username') or ''
    user_id = body.get('user_id')

    if not selected_empresa:
        return _json_http({'ok': False, 'message': 'Seleccione una empresa.'}, status=400)

    if not selected_periodo:
        return _json_http({'ok': False, 'message': 'Seleccione un periodo.'}, status=400)

    if not isinstance(rows, list) or not rows:
        return _json_http({'ok': False, 'message': 'No hay filas validadas para migrar.'}, status=400)

    try:
        batches = _chunk_asientos_rows(rows)
        total_batches = len(batches)
        total_inserted = 0
        first_inserted_nro = None
        last_inserted_nro = None
        migration_process = [
            'Migracion preparada en {0} lote(s) de hasta {1} asiento(s) por lote.'.format(
                total_batches,
                ASIENTOS_MIGRATION_BATCH_SIZE
            )
        ]

        for index, batch_rows in enumerate(batches, start=1):
            first_row = batch_rows[0] if batch_rows else {}
            lote_empresa = first_row.get('cod_empresa') or selected_empresa
            lote_nro = first_row.get('nrotransaccion') or ''
            migration_process.append(
                'Enviando lote {0} de {1} para Empresa {2} / NroTransaccion {3} con {4} fila(s).'.format(
                    index, total_batches, lote_empresa, lote_nro, len(batch_rows)
                )
            )

            try:
                resp = _node_post(
                    'migraciones/asientos/import',
                    json={
                        'empresa': selected_empresa,
                        'periodo': selected_periodo,
                        'rows': batch_rows,
                        'user_id': user_id,
                        'username': username,
                    },
                    timeout=NODE_API_TIMEOUT_LONG
                )
            except Exception as exc:
                migration_process.append('Fallo el lote {0} de {1}: {2}'.format(index, total_batches, str(exc)))
                return _json_http({
                    'ok': False,
                    'message': 'La migracion se interrumpio en el lote {0} de {1}.'.format(index, total_batches),
                    'process': migration_process,
                }, status=500)

            data = resp.json() if resp.content else {}
            batch_process = data.get('process') or []
            for step in batch_process:
                migration_process.append('[Lote {0}] {1}'.format(index, step))

            if resp.status_code >= 400 or not data.get('ok'):
                migration_process.append('El lote {0} devolvio error y se detuvo la migracion.'.format(index))
                return _json_http({
                    'ok': False,
                    'message': data.get('error') or 'No se pudo migrar el lote {0} de {1}.'.format(index, total_batches),
                    'process': migration_process,
                }, status=400)

            total_inserted += int(data.get('inserted_count') or 0)
            if data.get('from_nrotransac') and first_inserted_nro is None:
                first_inserted_nro = data.get('from_nrotransac')
            if data.get('to_nrotransac'):
                last_inserted_nro = data.get('to_nrotransac')
            migration_process.append('Lote {0} completado correctamente.'.format(index))

        migration_message = (
            'Migracion finalizada correctamente. Lotes procesados: {0}. Asientos insertados: {1}.'
        ).format(total_batches, total_inserted)
        if first_inserted_nro is not None and last_inserted_nro is not None:
            migration_message += ' Desde Asiento Nro.: {0} - Hasta Nro.: {1}.'.format(
                first_inserted_nro,
                last_inserted_nro
            )

        return _json_http({
            'ok': True,
            'message': migration_message,
            'process': migration_process,
            'inserted_count': total_inserted,
            'from_nrotransac': first_inserted_nro,
            'to_nrotransac': last_inserted_nro,
        })
    except Exception as exc:
        return _json_http({'ok': False, 'message': str(exc)}, status=500)


@csrf_exempt
@require_POST
def api_migraciones_ventas_validate(request):
    selected_empresa = request.POST.get('empresa') or ''
    selected_periodo = request.POST.get('periodo') or str(timezone.now().year)
    selected_tipo = request.POST.get('tipo_comprobante') or ''
    selected_sucursal = request.POST.get('sucursal') or ''
    uploaded_file = request.FILES.get('archivo_excel')

    if not uploaded_file:
        return _json_http({'ok': False, 'message': 'Adjunte un archivo Excel.'}, status=400)

    if selected_empresa:
        try:
            meta_resp = _node_get('empresas/{0}/importar_ventas'.format(selected_empresa))
            meta_resp.raise_for_status()
            meta_data = meta_resp.json().get('data', [])
            row = meta_data[0] if meta_data else {}
            importar_ventas = _normalize_string(
                row.get('importar_ventas', row.get('IMPORTAR_VENTAS', 'N'))
            ).upper()
            if importar_ventas != 'S':
                return _json_http({'ok': False, 'message': 'La empresa seleccionada no tiene habilitada la opcion Importar Ventas.'}, status=400)
        except Exception:
            return _json_http({'ok': False, 'message': 'No se pudo verificar si la empresa tiene habilitada la opcion Importar Ventas.'}, status=400)

    try:
        rows = _read_asientos_excel(uploaded_file)
        upload_result = _validate_ventas_rows(
            rows,
            selected_empresa,
            selected_periodo,
            selected_tipo,
            selected_sucursal,
        )
        summary = upload_result.get('summary', {}) if upload_result else {}
        validation_ok = summary.get('total_rows', 0) > 0 and summary.get('error_rows', 0) == 0

        return _json_http({
            'ok': True,
            'validation_ok': validation_ok,
            'message': (
                'Excel validado correctamente.'
                if validation_ok
                else 'Excel controlado con observaciones. Revise los errores antes de migrar.'
            ),
            'data': upload_result,
        })
    except Exception as exc:
        return _json_http({'ok': False, 'message': str(exc)}, status=400)


@csrf_exempt
@require_POST
def api_migraciones_ventas_migrate(request):
    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return _json_http({'ok': False, 'message': 'No se pudo leer la solicitud.'}, status=400)

    selected_empresa = body.get('empresa') or ''
    selected_periodo = body.get('periodo') or str(timezone.now().year)
    selected_tipo = body.get('tipo_comprobante') or ''
    selected_sucursal = body.get('sucursal') or ''
    selected_empresa_desc = body.get('empresa_desc') or ''
    rows = body.get('rows') or []
    username = body.get('username') or ''

    if not isinstance(rows, list) or not rows:
        return _json_http({'ok': False, 'message': 'No hay filas validadas para migrar.'}, status=400)

    try:
        resp = _node_post(
            'migraciones/ventas/import',
            json={
                'empresa': selected_empresa,
                'empresa_desc': selected_empresa_desc,
                'periodo': selected_periodo,
                'cod_tp_comp': selected_tipo,
                'cod_sucursal': selected_sucursal,
                'rows': rows,
                'username': username,
            },
            timeout=NODE_API_TIMEOUT_LONG
        )
        resp.raise_for_status()
        data = resp.json() or {}
        return _json_http({
            'ok': bool(data.get('ok')),
            'message': data.get('message') or data.get('error', ''),
            'process': data.get('process', []) or [],
        }, status=200 if data.get('ok') else 400)
    except Exception as exc:
        return _json_http({'ok': False, 'message': str(exc)}, status=500)



